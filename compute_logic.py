#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
compute_logic.py
--------------------------------------------------
核心计算逻辑，供后端调用。
支持：
1. 同时处理多个 Excel 文件 (openpyxl)。
2. 按 "Created Time" 日期字段进行过滤 (闭区间)。
3. 以 Seller SKU 为分组键输出各项指标。
"""

from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Dict, Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from metrics_core import classify_order_status, date_in_range, iter_order_rows, metric_rates


def compute_metrics(file_streams: Iterable[BytesIO], start_date: date, end_date: date):
    """核心接口：返回 (Workbook, stats_dict)"""
    stats: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    total_rows = 0

    for fs in file_streams:
        for seller_sku, sub, cancel, shipped, created in iter_order_rows(fs):
            if seller_sku is None:
                continue
            if not date_in_range(created, start_date, end_date):
                continue

            sku = str(seller_sku)
            s = stats[sku]
            s["total"] += 1
            total_rows += 1

            status_key = classify_order_status(sub, cancel, shipped)
            if status_key:
                s[status_key] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "订单指标"
    headers = [
        "Seller SKU", "订单数", "签收率(%)", "已完成率(%)", "已送达率(%)", "退款率(%)", "发货前取消率(%)", "发货后取消率(%)", "仍在途率(%)",
    ]
    ws.append(headers)

    for sku, m in sorted(stats.items(), key=lambda x: (-x[1]["total"], x[0])):
        total = m["total"]
        if total == 0:
            continue
        rates = metric_rates(m, total)

        ws.append([
            sku,
            total,
            rates["sign_rate"],
            rates["completed_rate"],
            rates["delivered_rate"],
            rates["refund_rate"],
            rates["cancel_before_rate"],
            rates["cancel_after_rate"],
            rates["in_transit_rate"],
        ])

    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 14

    return wb, stats 
