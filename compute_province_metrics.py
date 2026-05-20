#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按省份统计每个 SKU 的签收率和订单占比
------------------------------------------------
指标定义：
1. 订单数               = 该 SKU 在该省份的订单数
2. 已完成率(%)          = Order Substatus 为 "已完成" 或 "Completed" 且 Cancelation/Return Type 为空 的订单数 / 订单数 * 100
3. 已送达率(%)          = Order Substatus 为 "已送达" 或 "Delivered" / 订单数 * 100
4. 退款率(%)            = Order Substatus 含 "Return"/"Refund" / 订单数 * 100
5. 发货前取消率(%)      = Order Substatus 为 "已取消" 或 "Canceled" 且 Shipped Time 为空 / 订单数 * 100
6. 发货后取消率(%)      = Order Substatus 为 "已取消" 或 "Canceled" 且 Shipped Time 不为空 / 订单数 * 100
7. 仍在途率(%)          = Order Substatus 为 "运输中" 或 "In transit" / 订单数 * 100
8. 签收率(%)            = 已完成率 + 已送达率 + 退款率
9. 订单占比(%)          = 该 SKU 在此省份的订单数 / 该 SKU 总订单数 * 100

注意：
- 表格第二行是描述行，需要跳过。
- 仅依赖 openpyxl 进行 Excel 读写，避免只读取到第一列的问题。
"""

from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable
from datetime import date
from io import BytesIO

from metrics_core import (
    classify_order_status,
    date_in_range,
    iter_order_rows,
    locate_columns,
    metric_rates,
)
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

INPUT_FILE = "全部 订单-2025-07-08-21_50.xlsx"  # 如需处理其它文件，可修改此常量或传参
OUTPUT_FILE = "省份指标分析结果.xlsx"


def read_orders(file_path: Path):
    """读取 Excel，返回迭代器 (sku_id, province, substatus, cancel_type, shipped_time, created_time)"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active  # 默认第一个工作表

    # 正确读取标题行（read_only 模式下无法通过 ws[1] 获取完整行）
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header_row)
    col_indices = locate_columns(headers, include_province=True)

    # 从第3行开始遍历（第2行是描述行）
    for row in ws.iter_rows(min_row=3, values_only=True):
        seller_sku = row[col_indices["seller_sku"]]
        province = row[col_indices["province"]]
        substatus = row[col_indices["order_substatus"]]
        cancel_type = row[col_indices["cancel_type"]]
        shipped_time = row[col_indices["shipped_time"]]
        created_time = row[col_indices["created_time"]]
        yield seller_sku, province, substatus, cancel_type, shipped_time, created_time

    wb.close()

def compute_metrics(file_path: Path):
    """核心计算逻辑"""
    stats: Dict[str, Dict[str, Dict[str, int]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(int))
    )
    sku_totals: Dict[str, int] = defaultdict(int)

    total_rows = 0
    # read_orders 返回 6 个值（包含 created_time），此处一并解包，created 暂不使用
    for seller_sku, province, sub, cancel, shipped, created in read_orders(file_path):
        if seller_sku is None:
            continue  # 跳过无效行
        sku_id = str(seller_sku)
        prov = str(province).strip() if province is not None else ""
        total_rows += 1
        sku_totals[sku_id] += 1
        s = stats[sku_id][prov]
        s["total"] += 1

        status_key = classify_order_status(sub, cancel, shipped)
        if status_key:
            s[status_key] += 1

        # 其它状态直接忽略

    print(f"已读取 {total_rows} 行订单记录，发现 {len(stats)} 个 SKU")
    return stats, sku_totals


def compute_metrics_streams(file_streams: Iterable[BytesIO], start_date: date, end_date: date):
    stats: Dict[str, Dict[str, Dict[str, int]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    sku_totals: Dict[str, int] = defaultdict(int)
    for fs in file_streams:
        for seller_sku, province, sub, cancel, shipped, created in iter_order_rows(fs, include_province=True):
            if seller_sku is None:
                continue
            if not date_in_range(created, start_date, end_date):
                continue
            sku_id = str(seller_sku)
            prov = str(province).strip() if province is not None else ""
            sku_totals[sku_id] += 1
            s = stats[sku_id][prov]
            s["total"] += 1
            status_key = classify_order_status(sub, cancel, shipped)
            if status_key:
                s[status_key] += 1
    return stats, sku_totals


def build_result_workbook(
    stats: Dict[str, Dict[str, Dict[str, int]]], sku_totals: Dict[str, int]
) -> Workbook:
    """根据统计结果构建结果工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = "省份指标"

    headers = [
        "Seller SKU",
        "Province",
        "订单数",
        "订单占比(%)",
        "签收率(%)",
        "已完成率(%)",
        "已送达率(%)",
        "退款率(%)",
        "发货前取消率(%)",
        "发货后取消率(%)",
        "仍在途率(%)",
    ]
    ws.append(headers)

    # 写数据
    for sku, province_map in sorted(stats.items(), key=lambda x: x[0]):
        total_sku = sku_totals.get(sku, 0)
        for prov, m in sorted(province_map.items(), key=lambda x: (-x[1]["total"], x[0])):
            total = m["total"]
            rates = metric_rates(m, total)
            share_rate = total / total_sku * 100 if total_sku else 0

            ws.append([
                sku,
                prov,
                total,
                round(share_rate, 2),
                rates["sign_rate"],
                rates["completed_rate"],
                rates["delivered_rate"],
                rates["refund_rate"],
                rates["cancel_before_rate"],
                rates["cancel_after_rate"],
                rates["in_transit_rate"],
            ])

    # 自动调整列宽
    for col_idx, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 14

    return wb


def main():
    print("=== 订单指标计算程序 (openpyxl) ===")
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        raise FileNotFoundError(f"找不到输入文件: {INPUT_FILE}")

    stats, sku_totals = compute_metrics(input_path)
    if not stats:
        print("未找到任何可计算数据。")
        return

    wb = build_result_workbook(stats, sku_totals)
    wb.save(OUTPUT_FILE)
    print(f"计算完成，结果已保存为: {OUTPUT_FILE}")


if __name__ == "__main__":
    main() 
