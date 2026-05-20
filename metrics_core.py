#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Shared parsing and metric helpers for order analytics."""

from datetime import date, datetime
from typing import Dict, Iterable, List, Optional, Union
import csv
import re
from io import BytesIO, TextIOWrapper
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.utils.exceptions import InvalidFileException


BASE_COLUMNS = {
    "order_substatus": ["order substatus"],
    "cancel_type": ["cancelation/return type", "cancellation/return type"],
    "seller_sku": ["seller sku"],
    "shipped_time": ["shipped time"],
    "created_time": ["created time"],
}

PROVINCE_ALIASES = [
    "province",
    "state",
    "province/state",
    "state/province",
    "province name",
]

COMPLETED_STATUS = {"已完成", "completed"}
DELIVERED_STATUS = {"已送达", "delivered"}
CANCELED_STATUS = {"已取消", "canceled"}
IN_TRANSIT_STATUS = {"运输中", "in transit"}


def normalize(text) -> str:
    return str(text).strip().lower() if text is not None else ""


def locate_columns(headers: List[str], include_province: bool = False) -> Dict[str, int]:
    columns = dict(BASE_COLUMNS)
    if include_province:
        columns["province"] = PROVINCE_ALIASES

    header_map = {normalize(h): idx for idx, h in enumerate(headers) if h is not None}
    idx_map: Dict[str, int] = {}
    for key, aliases in columns.items():
        for alias in aliases:
            if alias in header_map:
                idx_map[key] = header_map[alias]
                break
        if key not in idx_map:
            raise KeyError(f"列缺失: {aliases[0]} (实际标题行: {headers})")
    return idx_map


def iter_order_rows(file_bytes: Union[str, bytes, BytesIO], include_province: bool = False):
    """Yield normalized row tuples from .xlsx or .csv inputs."""
    if isinstance(file_bytes, str):
        with open(file_bytes, "rb") as f:
            data = BytesIO(f.read())
    elif isinstance(file_bytes, bytes):
        data = BytesIO(file_bytes)
    else:
        data = file_bytes
        data.seek(0)

    try:
        wb = load_workbook(data, data_only=True, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        cols = locate_columns(list(header_row), include_province=include_province)

        for row in ws.iter_rows(min_row=3, values_only=True):
            values = [
                row[cols["seller_sku"]],
                row[cols["order_substatus"]],
                row[cols["cancel_type"]],
                row[cols["shipped_time"]],
                row[cols["created_time"]],
            ]
            if include_province:
                values.insert(1, row[cols["province"]])
            yield tuple(values)
        wb.close()
    except (InvalidFileException, BadZipFile):
        data.seek(0)
        wrapper = TextIOWrapper(data, encoding="utf-8-sig")
        reader = csv.reader(wrapper)
        headers = next(reader)
        cols = locate_columns(headers, include_province=include_province)
        next(reader, None)

        def cell(row, key):
            idx = cols[key]
            return row[idx] if idx < len(row) else None

        for row in reader:
            values = [
                cell(row, "seller_sku"),
                cell(row, "order_substatus"),
                cell(row, "cancel_type"),
                cell(row, "shipped_time"),
                cell(row, "created_time"),
            ]
            if include_province:
                values.insert(1, cell(row, "province"))
            yield tuple(values)


def to_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            return from_excel(val).date()
        except Exception:
            pass
    if isinstance(val, str):
        txt = val.strip().replace("年", "-").replace("月", "-").replace("日", "")
        patterns = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y",
        ]
        for fmt in patterns:
            try:
                return datetime.strptime(txt, fmt).date()
            except ValueError:
                continue
        match = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", txt)
        if match:
            try:
                return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
            except ValueError:
                pass
    return None


def date_in_range(value: Union[datetime, date, str, int, float, None], start: date, end: date) -> bool:
    parsed = to_date(value)
    return parsed is not None and start <= parsed <= end


def classify_order_status(substatus, cancel_type, shipped_time) -> Optional[str]:
    sub = normalize(substatus)
    cancel = normalize(cancel_type)
    shipped_empty = shipped_time is None or str(shipped_time).strip() == ""

    if sub in COMPLETED_STATUS and cancel == "":
        return "completed"
    if sub in DELIVERED_STATUS:
        return "delivered"
    if "return" in sub or "refund" in sub:
        return "refund"
    if sub in CANCELED_STATUS:
        return "cancel_before" if shipped_empty else "cancel_after"
    if sub in IN_TRANSIT_STATUS:
        return "in_transit"
    return None


def metric_rates(metrics: Dict[str, int], total_override: Optional[int] = None) -> Dict[str, float]:
    total = total_override if total_override is not None else metrics.get("total", 0)
    if not total:
        return {
            "sign_rate": 0,
            "completed_rate": 0,
            "delivered_rate": 0,
            "refund_rate": 0,
            "cancel_before_rate": 0,
            "cancel_after_rate": 0,
            "in_transit_rate": 0,
        }

    completed_rate = metrics.get("completed", 0) / total * 100
    delivered_rate = metrics.get("delivered", 0) / total * 100
    refund_rate = metrics.get("refund", 0) / total * 100
    return {
        "sign_rate": round(completed_rate + delivered_rate + refund_rate, 2),
        "completed_rate": round(completed_rate, 2),
        "delivered_rate": round(delivered_rate, 2),
        "refund_rate": round(refund_rate, 2),
        "cancel_before_rate": round(metrics.get("cancel_before", 0) / total * 100, 2),
        "cancel_after_rate": round(metrics.get("cancel_after", 0) / total * 100, 2),
        "in_transit_rate": round(metrics.get("in_transit", 0) / total * 100, 2),
    }

